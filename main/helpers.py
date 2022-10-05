from pathlib import Path

import openpyxl

from main.models import Product


def more_adding_warehouse_product(order):
    warehouses = []

    file = order.excel_product
    if file is not None:
        sheet = openpyxl.load_workbook(file.path, data_only=True).active
        max_row = sheet.max_row
        prev_obj_index = 2
        for i in range(2, max_row+1):
            if sheet.cell(row=i, column=1).value is not None:
                index = i
                prev_obj_index = i
            else:
                index = prev_obj_index

            warehouses.append(
                Product(
                    product_id=sheet.cell(row=index, column=1).value,
                    name=sheet.cell(row=index, column=2).value,
                    description=sheet.cell(row=index, column=3).value,
                    price=sheet.cell(row=index, column=4).value,
                    tags=sheet.cell(row=index, column=5).value,
                    image_url=sheet.cell(row=index, column=6).value,
                    option_Name=sheet.cell(row=index, column=7).value,
                    hidden=sheet.cell(row=index, column=8).value,
                    weight=sheet.cell(row=index, column=9).value,
                    hidden_btn = sheet.cell(row=index, column=10).value,
                    prepaymentPercent=sheet.cell(row=index, column=11).value,
                    bonusesPercent=sheet.cell(row=index, column=12).value
                )
            )

        Product.objects.bulk_create(warehouses)
    return warehouses[-1] if warehouses else None
